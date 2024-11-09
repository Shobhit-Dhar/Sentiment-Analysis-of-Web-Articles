import os
import nltk
from nltk.tokenize import word_tokenize,sent_tokenize
import string
import openpyxl

#nltk.download('cmudict')
#nltk.download('punkt')

# Load the CMU Pronouncing Dictionary
pronouncing_dict = nltk.corpus.cmudict.dict()

# personal pronoun list
personal_pronouns = ['i', 'me', 'my', 'mine', 'we', 'us', 'our', 'ours',
                     'you', 'your', 'yours', 'he', 'him', 'his', 'she',
                     'her', 'hers', 'it', 'its', 'they', 'them', 'their', 'theirs']

s1=set()
s2=set()
with open('positive-words.txt','r') as fp:
    s1.update(word.strip().lower() for word in fp.read().splitlines())
with open('negative-words.txt','r') as fn:
    s2.update(word.strip().lower() for word in fn.read().splitlines())

directory = os.getcwd() + r'\txtfiles'

def analyze_sentiment_positive(tokens, s1):
    sentiment_pscore = 0
    for word in tokens:
        if word in s1:
            sentiment_pscore += 1
    return sentiment_pscore

def analyze_sentiment_negative(tokens, s2):
    sentiment_nscore = 0
    for word in tokens:
        if word in s2:
            sentiment_nscore -= 1
    return sentiment_nscore*(-1)

# Function to remove punctuation from tokens
def remove_punctuation(tokens):
    return [word for word in tokens if word not in string.punctuation]

# Function to count the number of sentences
def count_sentences(text):
    sentences = sent_tokenize(text)
    return len(sentences)


# Function to count syllables in a word
def count_syllables(word):
    if word.lower() in pronouncing_dict:
        pronunciations = pronouncing_dict[word.lower()]
        # Take the first pronunciation and count vowels in the phoneme
        return len([phoneme for phoneme in pronunciations[0] if phoneme[-1].isdigit()])
    else:
        # If word is not in the dictionary, approximate by counting vowels
        return sum(1 for char in word if char in 'aeiouAEIOU')

# Function to count complex words (words with more than two syllables)
def count_complex_words(tokens):
    complex_word_count = 0
    cnt_syllables = 0
    for word in tokens:
        num_syllables = count_syllables(word)
        if num_syllables > 2:
            complex_word_count += 1
        cnt_syllables += num_syllables
    return complex_word_count,cnt_syllables

def count_personal_pronouns(tokens):
    num_personal_pronouns = 0
    for word in tokens:
        if word.lower() in personal_pronouns and word != "US":
            num_personal_pronouns+=1
    return num_personal_pronouns

def character_count(text):
    return len(text)

# Function to calculate average sentence length
def calculate_average_sentence_length(text):
    sentences = sent_tokenize(text)
    total_words = sum(len(word_tokenize(sentence)) for sentence in sentences)
    num_sentences = len(sentences)
    return total_words / num_sentences

def avg_num_words_per_sentence(word_count,num_sentence):
    return word_count/num_sentence


def process_file(file_path,excel_file_path):
    try:
        with open(file_path, 'r', encoding= 'utf-8') as file:
            text = file.read()

            # counting sentences
            num_sentence = count_sentences(text)

            #average sentence length
            avg_senlen = calculate_average_sentence_length(text)


            # tokenizing
            tokens = word_tokenize(text.lower())
            tokens = remove_punctuation(tokens)

            #word counts after removing punctuations
            word_count = len(tokens)

            # calculating average word length
            avg_word_length = character_count(text) / word_count

            # calculating avg words per sentence
            avg_wordpersen = avg_num_words_per_sentence(word_count,num_sentence)

            # Count complex words ------ return type tuple form ------ index 0 for complex words and index 1 for total syllable count
            cnt = count_complex_words(tokens)

            #percentage of complex words
            percent_complex_words = cnt[0]/word_count

            # calculating fog index
            fog_index = 0.4 * (avg_senlen + percent_complex_words)

            # syllable count per word
            num_syllable_per_word = cnt[1]/word_count

            #counting personal pronouns
            num_personal_pronouns = count_personal_pronouns(tokens)


            scorep = analyze_sentiment_positive(tokens, s1)
            scoren = analyze_sentiment_negative(tokens, s2)
            polsc = (scorep-scoren)/((scorep+scoren) + 0.000001)
            subscr = (scorep+scoren) / ((word_count) + 0.000001)



            '''print(f"Positive Sentiment Score = {scorep}")
            print(f"Negative Sentiment Score = {scoren}")
            print(f"Polarity score = {polsc}")
            print(f"Subjectivity score = {subscr}")
            print(f"Average sentence length = {avg_senlen}")
            print(f"Percentage of complex words = {percent_complex_words}")
            print(f"Fog Index = {fog_index}")
            print(f"Average number of words per sentence = {avg_wordpersen}")
            print(f"Complex word count = {cnt[0]}")
            print(f"Word count = {word_count}")
            print(f"Syllable count per word = {num_syllable_per_word}")
            print(f"Personal pronouns = {num_personal_pronouns}")
            print(f"Average word length = {avg_word_length}")'''

            # Write results to Excel file
            wb = openpyxl.load_workbook(excel_file_path)
            ws = wb.active

            # Write results to specific cells
            ws['C'+ str(i+2)] = scorep
            ws['D'+ str(i+2)] = scoren
            ws['E'+ str(i+2)] = polsc
            ws['F'+ str(i+2)] = subscr
            ws['G'+ str(i+2)] = avg_senlen
            ws['H'+ str(i+2)] = percent_complex_words
            ws['I'+ str(i+2)] = fog_index
            ws['J'+ str(i+2)] = avg_wordpersen
            ws['K'+ str(i+2)] = cnt[0]
            ws['L'+ str(i+2)] = word_count
            ws['M'+ str(i+2)] = num_syllable_per_word
            ws['N'+ str(i+2)] = num_personal_pronouns
            ws['O'+ str(i+2)] = avg_word_length

            # Save the Excel file
            wb.save(excel_file_path)

            print("Results saved to Excel file.")

    except FileNotFoundError:
        pass
    except UnicodeDecodeError:
        pass


# Load and tokenize text files


for i in range(100):
    file_name = 'blackassign000' + str(i + 1) + '.txt'
    file1_name = "Output Data Structure.xlsx"
    directory1 = os.getcwd()
    file_path = os.path.join(directory, file_name)
    excel_file_path = os.path.join(directory1,file1_name)
    texts = process_file(file_path,excel_file_path)




